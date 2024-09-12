from datetime import datetime
from  fake_useragent import UserAgent
from abc import ABC, abstractmethod
import threading
import time 
import os
import glob
import pandas as pd
import akshare as ak
from matplotlib import pyplot as plt
# import time
# import threading

class DataClass(ABC):
    '''
    base class:用来限定处理数据步骤
    '''
    def __init__(self,headers=None):
        # self._url = url
        self.headers = headers
        if self.headers is None:
            self.headers = {"User-Agent":getUserAgent()} 
    @abstractmethod
    def get_data_json(self,*args,**kwargs):
        pass
    
    def get_data_df(self,*args,**kwargs):
        pass

def getStrDate(mode):
    '''
    时间格式string生成
    '''
    match mode:
        case 0:
            return f"{datetime.today().strftime('%Y-%m-%d-%H-%M-%S')}"
        case 1:
            return f"{datetime.today().strftime('%Y%m%d')}"
        case _:
            return f"{datetime.today().strftime('%Y-%m-%d-%H-%M')}"

def getUserAgent():
    return UserAgent().random

def time_diplayer(func, *args, **kwargs):
    '''
    用来展示消耗时间的工具
    '''
    event = threading.Event()
    s = time.time()
    def timer(event,s):
        try:
            while not event.is_set():
                # current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                current_time = str(round(time.time() -s,3)) 
                print("\r" + current_time + " s", end='', flush=True)
                time.sleep(1)
        except Exception as e:
            raise e
    
    t = threading.Thread(target=timer, args=(event,s))
    t.start()
    try:
        ret = func(*args, **kwargs)
    except Exception as e:
        ret = None
        pass
    event.set()
    t.join()
    return ret


    
def get_code_name():
    import akshare as ak
    # 获取所有 A 股股票的实时行情数据
    stock_zh_a_spot_em = ak.stock_zh_a_spot_em()

    # 提取代码和名称列
    df = stock_zh_a_spot_em[["代码", "名称"]]

    # 重命名列
    df.columns = ["code", "name"]  
    return df,stock_zh_a_spot_em
    

def merge_excel_files(directory):
    # 指定包含Excel文件的目录
    # directory = 'path/to/excel/files'
    # 获取当前日期
    current_date = datetime.now().strftime('%Y-%m-%d')

    # 创建一个空的ExcelWriter对象
    with pd.ExcelWriter(f'merged_excel_{current_date}.xlsx', engine='openpyxl') as writer:
        # 遍历目录中的所有Excel文件
        for file_path in glob.glob(os.path.join(directory, '*.xlsx')):
            # 读取Excel文件
            df = pd.read_excel(file_path)
            
            # 获取文件名（不包括路径和扩展名）
            sheet_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # 将DataFrame写入Excel文件的新工作表中
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def merge_df_files(df_dict,prefix):
    '''
    df_dict:{"sheet_name":someDataframe,...}
    '''
    current_date = datetime.now().strftime('%Y-%m-%d')
    # 创建一个空的ExcelWriter对象
    dest = os.path
    with pd.ExcelWriter(prefix + f'_{current_date}.xlsx', engine='openpyxl') as writer:
        # 遍历
        for sheet_name,df in df_dict.items():
            # 获取文件名（不包括路径和扩展名）
            sheet_name = sheet_name
            # 将DataFrame写入Excel文件的新工作表中
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def add_charts(df):
    '''添加图片到excel中
    '''
    import pandas as pd
    import matplotlib.pyplot as plt
    from openpyxl import Workbook,load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
    import os

    # 生成走势图并保存为图片
    def plot_trend(code, trend_data):
        plt.figure(figsize=(10, 2))
        plt.plot(trend_data)
        plt.title(f'Control Trend for {code}')
        plt.xlabel('Time')
        plt.ylabel('Percentage')
        plt.grid(True)
        
        # 保存图片
        filename = f'{code}_trend.png'
        filepath = os.path.join('trends', filename)
        plt.savefig(filepath, bbox_inches='tight')
        plt.close()
        return filepath

    # 创建目录以保存图片
    if not os.path.exists('trends'):
        os.makedirs('trends')

    # 为每个 code 生成走势图
    df['Trend_Plot'] = df.apply(lambda row: plot_trend(row['代码'], row['近来控盘比例趋势']), axis=1)

    # 创建一个新的工作簿
    wb = Workbook()
    

    # 获取默认的工作表
    ws = wb.active

    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=str(value))

    # 添加图片到 Excel 文件
    for idx, row in df.iterrows():
        img = Image(row['Trend_Plot'])
        # 假设图片放置在每行的末尾
        col_img = len(df.columns) + 1
        row_img = idx + 2  # 行号，从第二行开始
        ws.add_image(img, f"{get_column_letter(col_img)}{row_img}")

    # 保存 Excel 文件
    output_file = 'output_with_trends.xlsx'
    wb.save(output_file)



def attention_kongpan(folder="trends"):
    import mplfinance as mpf
    from tqdm import tqdm
    import matplotlib.pyplot as plt
    import warnings
    warnings.filterwarnings("ignore")
    plt.rcParams['font.sans-serif']=['MicroSoft YaHei'] #用来正常显示中文标签
    plt.rcParams['axes.unicode_minus']=False #用来正常显示负号
    def kongpan_attention_kline_data():
        from ATTENTION import ATTENTION
        from utils import get_code_name
        code_name_df,spot_df = get_code_name()
        data = pd.DataFrame()
        data["代码"] = ATTENTION
        dfs = []
        for code in ATTENTION:
            stock_comment_detail_zlkp_jgcyd_em_df = ak.stock_comment_detail_zlkp_jgcyd_em(symbol=code)
            tmp_date = stock_comment_detail_zlkp_jgcyd_em_df["date"].map(lambda x:datetime.strftime(x,"%Y%m%d"))
            start_date = tmp_date[0]
            end_date = tmp_date.tolist()[-1]
            stock_zh_a_hist_df = ak.stock_zh_a_hist(symbol=code, 
                                                    period="daily", 
                                                    start_date=start_date, 
                                                    end_date=end_date, 
                                                    adjust="qfq")
            stock_comment_detail_zlkp_jgcyd_em_df.rename(columns={"value":"近来控盘比例趋势","date":"日期"},inplace=True)
            df = pd.merge(stock_zh_a_hist_df,stock_comment_detail_zlkp_jgcyd_em_df,on="日期")
            name = code_name_df[code_name_df["code"]==code]["name"].tolist()[0]
            df["名称"] = name
            # print(df)
            # trend.append([round(x,2) for x in stock_comment_detail_zlkp_jgcyd_em_df["value"].tolist()])
            dfs.append(df.copy())
        return dfs
    
    if not os.path.exists(folder):
            os.makedirs(folder)
    
    
    dfs = kongpan_attention_kline_data()
    
    for df in tqdm(dfs):
        # 修改列名
        df = df.rename(
            {
                "日期": "Date",
                "开盘": "Open",
                "收盘": "Close",
                "最低": "Low",
                "最高": "High",
                "成交量": "Volume",
            },
            axis=1,
        )

        # 将 Date 列设为索引
        df.index = df["Date"].astype("datetime64[ns]")
        df = df.sort_index()
        # df.set_index('Date', inplace=True)
        # 确保索引是 DatetimeIndex 类型
        # df.index = pd.DatetimeIndex(df.index)
        name = df["名称"].tolist()[0]
        #plot style
        mc = mpf.make_marketcolors(up='red',down='green',  volume={'up':'red','down':'green'})
        style = mpf.make_mpf_style(rc={'font.family': 'SimHei'},
                                    base_mpf_style= 'yahoo',
                                    marketcolors=mc)
        # 计算移动平均线
        df['MA5'] = df['Close'].rolling(window=5).mean()
        df['MA10'] = df['Close'].rolling(window=10).mean()
        df['MA20'] = df['Close'].rolling(window=20).mean()
        

        df["mean_kp"] =  df["近来控盘比例趋势"].mean() 
        df["std_kp"] =  df["近来控盘比例趋势"].std() 
        df["lowbound_kp"] = df["mean_kp"] - 3 * df["std_kp"]
        df["upperbound_kp"] = df["mean_kp"] + 3 * df["std_kp"]
        
        
        addplots = [
                    mpf.make_addplot(df["近来控盘比例趋势"], 
                                    color="b", 
                                    width=1,
                                    ylabel=" control trend ",
                                    y_on_right=True,
                                    panel=2,
                                    type="line",
                                    # secondary_y=True,
                                    ),
                    
                    mpf.make_addplot(df["mean_kp"] , 
                                    color="b", 
                                    width=1,
                                    # ylabel=" control trend ",
                                    panel=2,
                                    type="line",
                                    # scatter=True,
                                    linestyle="dashed",
                                    # secondary_y=True,
                                    ),
                     mpf.make_addplot(df["lowbound_kp"], 
                                    color="green", 
                                    width=1,
                                    # ylabel=" control trend ",
                                    panel=2,
                                    type="line",
                                    # scatter=True,
                                    linestyle="dashed",
                                    # secondary_y=True,
                                    ),
                      mpf.make_addplot(df["upperbound_kp"], 
                                    color="red", 
                                    width=1,
                                    # ylabel=" control trend ",
                                    panel=2,
                                    type="line",
                                    # scatter=True,
                                    linestyle="dashed",
                                    # secondary_y=True,
                                    ),
                    
                    
                    # mpf.make_addplot(df['MA5'], color='blue', width=1, type='line',),
                    # mpf.make_addplot(df['MA10'], color='orange', width=1, type='line'),
                    # mpf.make_addplot(df['MA20'], color='green', width=1, type='line',),
                    
                    
                    # mpf.make_addplot(df["upperbound_kp"], 
                    #                 color="red", 
                    #                 width=1,
                    #                 # ylabel=" control trend ",
                    #                 panel=3,
                    #                 type="line",
                    #                 # scatter=True,
                    #                 linestyle="dashed",
                    #                 # secondary_y=True,
                    #                 ),

                    ] 
        
        title = f'{name}_{datetime.now().strftime("%Y%m%d")}'
        kwargs = dict(
                    type='candle',
                    volume = True,
                    mav=(5,10,20),
                    scale_width_adjustment = dict(volume=0.5, candle=1.15,lines=0.65),
                    datetime_format='%m%d',
                    xrotation=90,
                    title=title,
                    ylabel='price',
                    ylabel_lower='volume',
                    style = style,
                    addplot=addplots,
                    # tight_layout=True,
                    figratio=(14, 8),
                    figscale=1.,
                    
                )
        # fig, axes = plt.subplots(nrows=3, ncols=1, figsize=(10, 8))
        fig,axes = mpf.plot(df,returnfig=True,**kwargs)
        
        #定制一下
        lines = [plt.Line2D([0], [0], color=color, lw=2) for color in ['blue', 'orange', 'green']]
        labels = ['MA5', 'MA10', 'MA20']
        axes[0].legend(lines, labels, loc='upper left')
        
        
        #最高最低3点
        # 找出最低和最高的3个点
        lowest_points = df.nsmallest(3, '近来控盘比例趋势')
        highest_points = df.nlargest(3, '近来控盘比例趋势')
        # 准备散点数据
        scatter_lowest = lowest_points[['近来控盘比例趋势']]
        scatter_highest = highest_points[['近来控盘比例趋势']]
        
       
        id_ = 4  # The index of the panel where control trend is plotted
        axes[id_].scatter(df.index.strftime('%m%d'),df["近来控盘比例趋势"],color="b",marker="o",s=2)
        axes[id_].scatter(scatter_lowest.index.strftime('%m%d'),scatter_lowest["近来控盘比例趋势"],color="green",marker="s",s=10)
        axes[id_].scatter(scatter_highest.index.strftime('%m%d'),scatter_highest["近来控盘比例趋势"],color="red",marker="s",s=10)
        
        for idx, row in scatter_lowest.iterrows():
            date = idx.strftime('%m%d')
            value = row['近来控盘比例趋势']
            axes[id_].annotate(f'{value:.2f}%', (date, value), textcoords="offset points", xytext=(5, 10), ha='center')
        for idx, row in scatter_highest.iterrows():
            date = idx.strftime('%m%d')
            value = row['近来控盘比例趋势']
            axes[id_].annotate(f'{value:.2f}%', (date, value), textcoords="offset points", xytext=(5, -10), ha='center')

                
        
        #
        # ax_new = fig.add_subplot(414)
        # ax_new.plot()
        # ax_new.scatter(scatter_lowest.index.strftime('%m%d'),scatter_lowest["近来控盘比例趋势"],color="green",marker="s",s=10)
        # ax_new.scatter(scatter_highest.index.strftime('%m%d'),scatter_highest["近来控盘比例趋势"],color="red",marker="s",s=10)
        
        # fig.tight_layout()
        
        fig.savefig(f'{folder}/{title}.png',format="png")


def attention_kongpan2(folder="trends"):
    import mplfinance as mpf
    from tqdm import tqdm
    import matplotlib.pyplot as plt
    import warnings
    warnings.filterwarnings("ignore")
    plt.rcParams['font.sans-serif']=['MicroSoft YaHei'] #用来正常显示中文标签
    plt.rcParams['axes.unicode_minus']=False #用来正常显示负号
    def kongpan_attention_kline_data():
        from ATTENTION import ATTENTION
        from utils import get_code_name
        code_name_df,spot_df = get_code_name()
        data = pd.DataFrame()
        data["代码"] = ATTENTION
        dfs = []
        for code in ATTENTION:
            stock_comment_detail_zlkp_jgcyd_em_df = ak.stock_comment_detail_zlkp_jgcyd_em(symbol=code)
            tmp_date = stock_comment_detail_zlkp_jgcyd_em_df["date"].map(lambda x:datetime.strftime(x,"%Y%m%d"))
            start_date = tmp_date[0]
            end_date = tmp_date.tolist()[-1]
            stock_zh_a_hist_df = ak.stock_zh_a_hist(symbol=code, 
                                                    period="daily", 
                                                    start_date=start_date, 
                                                    end_date=end_date, 
                                                    adjust="qfq")
            stock_comment_detail_zlkp_jgcyd_em_df.rename(columns={"value":"近来控盘比例趋势","date":"日期"},inplace=True)
            df = pd.merge(stock_zh_a_hist_df,stock_comment_detail_zlkp_jgcyd_em_df,on="日期")
            name = code_name_df[code_name_df["code"]==code]["name"].tolist()[0]
            df["名称"] = name
            # print(df)
            # trend.append([round(x,2) for x in stock_comment_detail_zlkp_jgcyd_em_df["value"].tolist()])
            dfs.append(df.copy())
        return dfs
    
    if not os.path.exists(folder):
        os.makedirs(folder)
    
    
    dfs = kongpan_attention_kline_data()
    
    for df in tqdm(dfs):
        # 修改列名
        df = df.rename(
            {
                "日期": "Date",
                "开盘": "Open",
                "收盘": "Close",
                "最低": "Low",
                "最高": "High",
                "成交量": "Volume",
            },
            axis=1,
        )

        # 将 Date 列设为索引
        df.index = df["Date"].astype("datetime64[ns]")
        df = df.sort_index()
        name = df["名称"].tolist()[0]
        #plot style
        mc = mpf.make_marketcolors(up='red',down='green',  volume={'up':'red','down':'green'})
        style = mpf.make_mpf_style(rc={'font.family': 'SimHei'},
                                    base_mpf_style= 'yahoo',
                                    marketcolors=mc)
        # 计算移动平均线
        df['MA5'] = df['Close'].rolling(window=5).mean()
        df['MA10'] = df['Close'].rolling(window=10).mean()
        df['MA20'] = df['Close'].rolling(window=20).mean()
        

        df["mean_kp"] =  df["近来控盘比例趋势"].mean() 
        df["std_kp"] =  df["近来控盘比例趋势"].std() 
        df["lowbound_kp"] = df["mean_kp"] - 3 * df["std_kp"]
        df["upperbound_kp"] = df["mean_kp"] + 3 * df["std_kp"]
        
        
        
        # 创建一个 Figure 和多个子图
        fig, axes = plt.subplots(3,
                                 1, 
                                figsize=(14, 8), 
                                gridspec_kw={'height_ratios': [3, 1, 1]}, 
                                sharex=True)

        title = f'{name}_{datetime.now().strftime("%Y%m%d")}'
        # fig.suptitle(title)
        
        # 使用 mplfinance 绘制 K 线图
        mpf_fig = mpf.plot(df, type='candle', 
                 ax=axes[0], 
                 volume=False, 
                #  mav=(5, 10, 20),
                style=style, 
                datetime_format='%m%d', 
                xrotation=90, 
                title=name,
                ylabel='price',)
        
        # 设置图例
        axes[0].legend(loc='upper left')
        
        
        # 使用 matplotlib 绘制交易量
        colors = df['Close'] > df['Open']
        colors = colors.map({True: 'red', False: 'green'})
        axes[1].bar(df.index, df['Volume'], color=colors, alpha=0.4)
        axes[1].set_ylabel('Volume')
        
        
        # 使用 matplotlib 绘制控盘比例趋势
        axes[2].plot(df.index.strftime('%m%d'), df["近来控盘比例趋势"], color="b", label="Control Trend")
        axes[2].plot(df.index.strftime('%m%d'), df["mean_kp"], color="b", linestyle="dashed", label="Mean")
        axes[2].plot(df.index.strftime('%m%d'), df["lowbound_kp"], color="green", linestyle="dashed", label="Lower Bound")
        axes[2].plot(df.index.strftime('%m%d'), df["upperbound_kp"], color="red", linestyle="dashed", label="Upper Bound")

        
        
        
        
        # #定制一下
        # lines = [plt.Line2D([0], [0], color=color, lw=2) for color in ['blue', 'orange', 'green']]
        # labels = ['MA5', 'MA10', 'MA20']
        # axes[0].legend(lines, labels, loc='upper left')
        
        
        #最高最低3点
        # 找出最低和最高的3个点
        lowest_points = df.nsmallest(3, '近来控盘比例趋势')
        highest_points = df.nlargest(3, '近来控盘比例趋势')
        # 准备散点数据
        scatter_lowest = lowest_points[['近来控盘比例趋势']]
        scatter_highest = highest_points[['近来控盘比例趋势']]
        
       
        id_ = 2  # The index of the panel where control trend is plotted
        axes[2].scatter(df.index.strftime('%m%d'),df["近来控盘比例趋势"],color="b",marker="o",s=2)
        axes[2].scatter(scatter_lowest.index.strftime('%m%d'),scatter_lowest["近来控盘比例趋势"],color="green",marker="s",s=10)
        axes[2].scatter(scatter_highest.index.strftime('%m%d'),scatter_highest["近来控盘比例趋势"],color="red",marker="s",s=10)
        
        for idx, row in scatter_lowest.iterrows():
            date = idx.strftime('%m%d')
            value = row['近来控盘比例趋势']
            axes[2].annotate(f'{value:.2f}%', (date, value), textcoords="offset points", xytext=(5, 10), ha='center')
        for idx, row in scatter_highest.iterrows():
            date = idx.strftime('%m%d')
            value = row['近来控盘比例趋势']
            axes[2].annotate(f'{value:.2f}%', (date, value), textcoords="offset points", xytext=(5, -10), ha='center')

                
        
        fig.savefig(f'{folder}/{title}.png',format="png")


def closest_trade_date():
    # from instock.crawling.trade_date_hist import tool_trade_date_hist_sina
    import akshare as ak
    tool_trade_date_hist_df = ak.tool_trade_date_hist_sina()
    # print(tool_trade_date_hist_df)
    from datetime import datetime
    # print(datetime.now().strftime("%Y-%m-%d"))
    t = tool_trade_date_hist_df[tool_trade_date_hist_df["trade_date"] <= datetime.now().date()].iloc[-1].values[0].strftime("%Y%m%d")
    # print(t)
    return t



def is_now_open():
    from datetime import datetime,time
    now = datetime.now()
    trade_closest_date = closest_trade_date()
    current_date = now.date().strftime("%Y%m%d")
    if current_date > trade_closest_date:
        return False
    else:
        start_time = datetime.combine(now.date(), time(9, 30))  # 构造今天的9点半时间  
        end_time = datetime.combine(now.date(), time(15, 0))    # 构造今天的15点时间  
        return start_time <= now <= end_time 

def is_now_break():
    from datetime import datetime,time
    now = datetime.now()
    start_time = datetime.combine(now.date(), time(11, 31))  # 构造今天的9点半时间  
    end_time = datetime.combine(now.date(), time(12, 59))    # 构造今天的15点时间  
    return start_time <= now <= end_time

 
if __name__ == "__main__":
    
        # attention_kongpan(folder="trends")
        print(is_now_open())
        # mpf.show()
        # pass


